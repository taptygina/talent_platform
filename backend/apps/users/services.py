import re
import secrets
import string
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from django.contrib.auth import get_user_model
from django.db import transaction
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas

from apps.users.models import UserRole

User = get_user_model()

REQUIRED_COLUMNS = {"first_name", "last_name"}
OPTIONAL_COLUMNS = {"middle_name", "email", "phone", "group_name"}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS
TEMPLATE_COLUMNS = ("first_name", "last_name", "middle_name", "email", "phone", "group_name")
USERNAME_MIN_LEN = 3
USERNAME_MAX_LEN = 30
PASSWORD_MIN_LEN = 8


def _normalize_text_key(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip()).casefold()


def _normalize_email_key(value: str) -> str:
    return (value or "").strip().casefold()


def _normalize_phone_key(value: str) -> str:
    return re.sub(r"[^\d+]+", "", (value or "").strip())


def _person_dedupe_key(
    *,
    role: str,
    first_name: str,
    last_name: str,
    middle_name: str,
    group_name: str,
) -> tuple[str, str, str, str, str]:
    return (
        role,
        _normalize_text_key(last_name),
        _normalize_text_key(first_name),
        _normalize_text_key(middle_name),
        _normalize_text_key(group_name) if role == UserRole.STUDENT else "",
    )


def _resolve_pdf_font_names() -> tuple[str, str]:
    """
    Register a Unicode font for Cyrillic rendering.
    """
    candidates = [
        ("DejaVuSans", Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"), "DejaVuSansBold", Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf")),
        ("DejaVuSans", Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")),
        ("LiberationSans", Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf")),
        ("NotoSans", Path("/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf")),
        ("DejaVuSans", Path("C:/Windows/Fonts/DejaVuSans.ttf")),
        ("Arial", Path("C:/Windows/Fonts/arial.ttf")),
        ("Tahoma", Path("C:/Windows/Fonts/tahoma.ttf")),
    ]
    for item in candidates:
        try:
            if len(item) == 4:
                font_name, font_path, bold_name, bold_path = item
                if font_path.exists() and bold_path.exists():
                    if font_name not in pdfmetrics.getRegisteredFontNames():
                        pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
                    if bold_name not in pdfmetrics.getRegisteredFontNames():
                        pdfmetrics.registerFont(TTFont(bold_name, str(bold_path)))
                    return font_name, bold_name
            else:
                font_name, font_path = item
                if font_path.exists():
                    if font_name not in pdfmetrics.getRegisteredFontNames():
                        pdfmetrics.registerFont(TTFont(font_name, str(font_path)))
                    return font_name, font_name
        except Exception:
            continue
    raise ValueError("РќРµ РЅР°Р№РґРµРЅ С€СЂРёС„С‚ СЃ РїРѕРґРґРµСЂР¶РєРѕР№ РєРёСЂРёР»Р»РёС†С‹ РґР»СЏ РіРµРЅРµСЂР°С†РёРё PDF.")


@dataclass
class ImportResult:
    created: int
    skipped: int
    generated_accounts: list[dict]
    errors: list[str]


def _normalize_username(value: str) -> str:
    value = value.lower()
    value = re.sub(r"[^a-z0-9._-]+", "", value)
    return value or "user"


def _username_seed(value: str) -> str:
    raw = re.sub(r"[^a-z0-9]+", "", _normalize_username(value))
    if len(raw) < USERNAME_MIN_LEN:
        raw = f"{raw}{secrets.token_hex(2)}"
    return raw[:USERNAME_MAX_LEN]


def _unique_username(base: str) -> str:
    seed = _username_seed(base)
    if not User.objects.filter(username=seed).exists():
        return seed

    attempts = 0
    while attempts < 500:
        tail_len = max(1, USERNAME_MAX_LEN - USERNAME_MIN_LEN)
        suffix = "".join(secrets.choice(string.digits) for _ in range(tail_len))
        candidate = f"{seed[:USERNAME_MIN_LEN]}{suffix}"[:USERNAME_MAX_LEN]
        if USERNAME_MIN_LEN <= len(candidate) <= USERNAME_MAX_LEN and not User.objects.filter(username=candidate).exists():
            return candidate
        attempts += 1

    raise ValueError(f"Не удалось сгенерировать уникальный логин длиной {USERNAME_MIN_LEN}-{USERNAME_MAX_LEN} символов.")


def _generate_password(length: int = 10) -> str:
    if length < PASSWORD_MIN_LEN:
        raise ValueError(f"РџР°СЂРѕР»СЊ РґРѕР»Р¶РµРЅ СЃРѕРґРµСЂР¶Р°С‚СЊ РјРёРЅРёРјСѓРј {PASSWORD_MIN_LEN} СЃРёРјРІРѕР»РѕРІ.")
    alphabet = string.ascii_letters + string.digits
    return "".join(secrets.choice(alphabet) for _ in range(length))


def _cell_text(row: dict, key: str) -> str:
    value = row.get(key)
    return str(value).strip() if value is not None else ""


def build_import_template_xlsx() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "users-import"

    for index, column in enumerate(TEMPLATE_COLUMNS, start=1):
        sheet.cell(row=1, column=index, value=column)

    sample_values = ("РРІР°РЅ", "РРІР°РЅРѕРІ", "РРІР°РЅРѕРІРёС‡", "ivanov@example.com", "+79990000000", "РРЎ-222Р±")
    for index, value in enumerate(sample_values, start=1):
        sheet.cell(row=2, column=index, value=value)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def build_credentials_pdf(accounts: list[dict], role: str) -> bytes:
    output = BytesIO()
    pdf = canvas.Canvas(output, pagesize=A4)
    width, height = A4
    font_name, font_name_bold = _resolve_pdf_font_names()

    y = height - 40
    pdf.setFont(font_name_bold, 13)
    pdf.drawString(40, y, "РЎРіРµРЅРµСЂРёСЂРѕРІР°РЅРЅС‹Рµ СѓС‡РµС‚РЅС‹Рµ РґР°РЅРЅС‹Рµ")
    y -= 18
    pdf.setFont(font_name, 10)
    pdf.drawString(40, y, f"Р РѕР»СЊ: {role}")
    y -= 24

    pdf.setFont(font_name, 9)
    for index, account in enumerate(accounts, start=1):
        needed_space = 9 * 8 + 20
        if y < needed_space:
            pdf.showPage()
            y = height - 40
            pdf.setFont(font_name, 9)

        pdf.setFont(font_name_bold, 9)
        pdf.drawString(40, y, f"{index}. {account.get('last_name', '')} {account.get('first_name', '')} {account.get('middle_name', '')}".strip())
        y -= 12
        pdf.setFont(font_name, 9)
        pdf.drawString(50, y, f"ID: {account.get('id', '')}")
        y -= 10
        pdf.drawString(50, y, f"Р›РѕРіРёРЅ: {account.get('username', '')}")
        y -= 10
        pdf.drawString(50, y, f"РџР°СЂРѕР»СЊ: {account.get('password', '')}")
        y -= 10
        pdf.drawString(50, y, f"Email: {account.get('email', '') or '-'}")
        y -= 10
        pdf.drawString(50, y, f"РўРµР»РµС„РѕРЅ: {account.get('phone', '') or '-'}")
        y -= 10
        pdf.drawString(50, y, f"Р“СЂСѓРїРїР°: {account.get('group_name', '') or '-'}")
        y -= 10
        pdf.drawString(50, y, f"Р РѕР»СЊ: {account.get('role', role)}")
        y -= 12
        pdf.line(40, y, width - 40, y)
        y -= 12

    pdf.save()
    return output.getvalue()


@transaction.atomic
def import_users_from_xlsx(file_bytes: bytes, role: str) -> ImportResult:
    if role not in UserRole.values:
        raise ValueError("РќРµРїРѕРґРґРµСЂР¶РёРІР°РµРјР°СЏ СЂРѕР»СЊ")

    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel-С„Р°Р№Р» РїСѓСЃС‚")

    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    header_map = {column: idx for idx, column in enumerate(header) if column}

    missing = sorted(REQUIRED_COLUMNS - set(header_map.keys()))
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    created = 0
    skipped = 0
    errors: list[str] = []
    generated_accounts: list[dict] = []
    seen_person_keys: set[tuple[str, str, str, str, str]] = set()

    existing_person_keys = {
        _person_dedupe_key(
            role=user.role,
            first_name=user.first_name,
            last_name=user.last_name,
            middle_name=user.middle_name,
            group_name=user.group_name,
        )
        for user in User.objects.filter(role=role).only("role", "first_name", "last_name", "middle_name", "group_name")
    }
    existing_email_keys = {
        _normalize_email_key(user.email)
        for user in User.objects.exclude(email="")
        .only("email")
        if _normalize_email_key(user.email)
    }
    existing_phone_keys = {
        _normalize_phone_key(user.phone)
        for user in User.objects.exclude(phone="")
        .only("phone")
        if _normalize_phone_key(user.phone)
    }

    for row_index, values in enumerate(rows[1:], start=2):
        row_data = {name: values[idx] for name, idx in header_map.items() if idx < len(values)}
        first_name = _cell_text(row_data, "first_name")
        last_name = _cell_text(row_data, "last_name")
        middle_name = _cell_text(row_data, "middle_name")
        email = _cell_text(row_data, "email")
        phone = _cell_text(row_data, "phone")
        group_name = _cell_text(row_data, "group_name")

        if not first_name or not last_name:
            skipped += 1
            errors.append(f"РЎС‚СЂРѕРєР° {row_index}: first_name Рё last_name РѕР±СЏР·Р°С‚РµР»СЊРЅС‹")
            continue

        person_key = _person_dedupe_key(
            role=role,
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            group_name=group_name,
        )
        email_key = _normalize_email_key(email)
        phone_key = _normalize_phone_key(phone)

        if person_key in seen_person_keys:
            skipped += 1
            errors.append(f"РЎС‚СЂРѕРєР° {row_index}: РґСѓР±Р»РёСЂСѓРµС‚СЃСЏ РІ Р·Р°РіСЂСѓР¶РµРЅРЅРѕРј С„Р°Р№Р»Рµ (Р¤РРћ/РіСЂСѓРїРїР°).")
            continue
        if person_key in existing_person_keys:
            skipped += 1
            errors.append(f"РЎС‚СЂРѕРєР° {row_index}: РїРѕР»СЊР·РѕРІР°С‚РµР»СЊ СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚ (Р¤РРћ/РіСЂСѓРїРїР°).")
            continue
        if email_key and email_key in existing_email_keys:
            skipped += 1
            errors.append(f"РЎС‚СЂРѕРєР° {row_index}: РїРѕР»СЊР·РѕРІР°С‚РµР»СЊ СЃ С‚Р°РєРёРј email СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚.")
            continue
        if phone_key and phone_key in existing_phone_keys:
            skipped += 1
            errors.append(f"РЎС‚СЂРѕРєР° {row_index}: РїРѕР»СЊР·РѕРІР°С‚РµР»СЊ СЃ С‚Р°РєРёРј С‚РµР»РµС„РѕРЅРѕРј СѓР¶Рµ СЃСѓС‰РµСЃС‚РІСѓРµС‚.")
            continue

        base = _normalize_username(f"{last_name}.{first_name}")
        try:
            username = _unique_username(base)
            password = _generate_password()
        except ValueError as exc:
            skipped += 1
            errors.append(f"РЎС‚СЂРѕРєР° {row_index}: {exc}")
            continue

        user_data = {
            "username": username,
            "password": password,
            "role": role,
            "first_name": first_name,
            "last_name": last_name,
            "middle_name": middle_name,
            "email": email,
            "phone": phone,
            "group_name": group_name,
        }
        user = User.objects.create_user(**user_data)
        created += 1
        seen_person_keys.add(person_key)
        existing_person_keys.add(person_key)
        if email_key:
            existing_email_keys.add(email_key)
        if phone_key:
            existing_phone_keys.add(phone_key)
        generated_accounts.append(
            {
                "id": user.id,
                "username": user.username,
                "password": password,
                "first_name": user.first_name,
                "last_name": user.last_name,
                "middle_name": user.middle_name,
                "full_name": user.full_name,
                "email": user.email,
                "phone": user.phone,
                "group_name": user.group_name,
                "role": user.role,
            }
        )

    return ImportResult(
        created=created,
        skipped=skipped,
        generated_accounts=generated_accounts,
        errors=errors,
    )
