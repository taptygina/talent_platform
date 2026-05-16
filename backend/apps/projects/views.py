from datetime import date, timedelta
import logging
from io import BytesIO
from django.db.models import BooleanField, Count, Exists, OuterRef, Q, Value
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.utils.dateparse import parse_date
from django.utils import timezone
from rest_framework import mixins, permissions, status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, MultiPartParser
from rest_framework.response import Response
from django.conf import settings

from apps.common.api_errors import error_response
from apps.file_security import FileValidationError, UploadPolicy, sanitize_filename, validate_uploaded_file
from apps.notifications.models import NotificationType
from apps.notifications.services import create_notifications
from apps.projects.api_messages import MESSAGES
from apps.projects.filters import ProjectFilter
from apps.projects.models import (
    Project,
    ProjectComment,
    ProjectLike,
    ProjectStage,
    ProjectStatus,
    ProjectType,
    ProjectSupervisorInvite,
    StageStatus,
    SupervisorInviteStatus,
    Team,
)
from apps.projects.permissions import IsCuratorOrReadOnly, IsTeacherCuratorOrReadOnly
from apps.projects.serializers import (
    ProjectCommentSerializer,
    ProjectDetailSerializer,
    ProjectLikeSerializer,
    ProjectListSerializer,
    ProjectSupervisorInviteSerializer,
    ProjectStageSerializer,
    TeamManageSerializer,
    TeamSerializer,
)
from apps.projects.tasks import dispatch_comment_notifications_task
from apps.users.models import User, UserRole
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)
MAX_IMAGE_UPLOAD_SIZE = 8 * 1024 * 1024


def _month_start(year: int, month: int) -> date:
    return date(year=year, month=month, day=1)


def _shift_month(source: date, delta_months: int) -> date:
    month_index = source.year * 12 + (source.month - 1) + delta_months
    year = month_index // 12
    month = month_index % 12 + 1
    return _month_start(year, month)


def _month_range(year: int, month: int) -> tuple[date, date]:
    start = _month_start(year, month)
    end = _shift_month(start, 1) - timedelta(days=1)
    return start, end


def _load_zone(active_projects: int, low_max: int, normal_max: int, high_max: int) -> str:
    if active_projects <= low_max:
        return "low"
    if active_projects <= normal_max:
        return "normal"
    if active_projects <= high_max:
        return "high"
    return "critical"


class ProjectViewSet(viewsets.ModelViewSet):
    filterset_class = ProjectFilter
    search_fields = ("title", "description", "goal", "supervisor__first_name", "supervisor__last_name", "team__name")
    ordering_fields = ("created_at", "start_date", "end_date", "title", "status")

    def get_queryset(self):
        # РЎС‡РёС‚Р°РµРј Р°РіСЂРµРіР°С‚С‹ Р·Р°СЂР°РЅРµРµ, С‡С‚РѕР±С‹ РёР·Р±РµР¶Р°С‚СЊ Р»РёС€РЅРёС… РґРѕРїРѕР»РЅРёС‚РµР»СЊРЅС‹С… Р·Р°РїСЂРѕСЃРѕРІ РІ СЃРїРёСЃРєР°С… Рё РєР°СЂС‚РѕС‡РєР°С….
        queryset = (
            Project.objects.select_related("supervisor", "team")
            .prefetch_related("participants", "team__members")
            .annotate(
                participants_count=Count("participants", distinct=True),
                likes_count=Count("likes", distinct=True),
                comments_count=Count("comments", distinct=True),
            )
        )
        if self.request.user.is_authenticated:
            queryset = queryset.annotate(
                liked_by_me=Exists(ProjectLike.objects.filter(project_id=OuterRef("pk"), user_id=self.request.user.id))
            )
        else:
            queryset = queryset.annotate(liked_by_me=Value(False, output_field=BooleanField()))
        if self.action == "list":
            # РђСЂС…РёРІ РїРѕ СѓРјРѕР»С‡Р°РЅРёСЋ СЃРєСЂС‹С‚ Рё РїРѕРєР°Р·С‹РІР°РµС‚СЃСЏ С‚РѕР»СЊРєРѕ РїРѕ СЏРІРЅРѕРјСѓ С„Р»Р°РіСѓ.
            include_archived = (self.request.query_params.get("include_archived") or "").strip().lower()
            if include_archived not in {"1", "true", "yes"}:
                queryset = queryset.filter(is_archived=False)
        return queryset

    def get_serializer_class(self):
        if self.action == "list":
            return ProjectListSerializer
        return ProjectDetailSerializer

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy"}:
            return [IsTeacherCuratorOrReadOnly()]
        return [permissions.IsAuthenticated()]

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def groups(self, request):
        search = (request.query_params.get("search") or "").strip()
        queryset = User.objects.filter(role=UserRole.STUDENT)
        if search:
            queryset = queryset.filter(group_name__icontains=search)
        groups = (
            queryset
            .exclude(group_name="")
            .values("group_name")
            .annotate(students_count=Count("id"))
            .order_by("group_name")
        )[:100]
        return Response(list(groups))

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def teams(self, request):
        queryset = Team.objects.exclude(name__startswith="group:").prefetch_related("members").select_related("supervisor")
        serializer = TeamSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def students(self, request):
        search = (request.query_params.get("search") or "").strip()
        group_name = (request.query_params.get("group_name") or "").strip()
        queryset = User.objects.filter(role=UserRole.STUDENT)
        if group_name:
            queryset = queryset.filter(group_name=group_name)
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(group_name__icontains=search)
                | Q(username__icontains=search)
            )
        students = (
            queryset.order_by("last_name", "first_name")
            .values("id", "username", "first_name", "last_name", "group_name")[:100]
        )
        return Response(list(students))

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def teachers(self, request):
        search = (request.query_params.get("search") or "").strip()
        queryset = User.objects.filter(role=UserRole.TEACHER)
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) | Q(last_name__icontains=search) | Q(username__icontains=search)
            )
        teachers = (
            queryset.order_by("last_name", "first_name")
            .values("id", "username", "first_name", "last_name", "email")[:100]
        )
        return Response(list(teachers))

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def like(self, request, pk=None):
        project = self.get_object()
        _, created = ProjectLike.objects.get_or_create(project=project, user=request.user)
        return Response({"created": created}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def unlike(self, request, pk=None):
        project = self.get_object()
        ProjectLike.objects.filter(project=project, user=request.user).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def publish(self, request, pk=None):
        project = self.get_object()
        if request.user.role not in {UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if project.status != ProjectStatus.DONE:
            return Response(
                {"detail": "РџСЂРѕРµРєС‚ РјРѕР¶РЅРѕ РїСѓР±Р»РёРєРѕРІР°С‚СЊ С‚РѕР»СЊРєРѕ СЃРѕ СЃС‚Р°С‚СѓСЃРѕРј В«Р—Р°РІРµСЂС€РµРЅВ»."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if not project.cover_image_url:
            return Response(
                {"detail": "Р”Р»СЏ РїСѓР±Р»РёРєР°С†РёРё С‚СЂРµР±СѓРµС‚СЃСЏ РѕР±Р»РѕР¶РєР° РїСЂРѕРµРєС‚Р°."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        project.is_published = True
        project.save(update_fields=["is_published", "updated_at"])
        return Response({"detail": "РџСЂРѕРµРєС‚ РѕРїСѓР±Р»РёРєРѕРІР°РЅ."})

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def unpublish(self, request, pk=None):
        project = self.get_object()
        if request.user.role not in {UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        project.is_published = False
        project.save(update_fields=["is_published", "updated_at"])
        return Response({"detail": "РџСѓР±Р»РёРєР°С†РёСЏ РїСЂРѕРµРєС‚Р° СЃРЅСЏС‚Р°."})


    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def archive(self, request, pk=None):
        project = self.get_object()
        if request.user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if request.user.role == UserRole.TEACHER and project.supervisor_id != request.user.id:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        project.is_archived = True
        project.save(update_fields=["is_archived", "updated_at"])
        return Response({"detail": "РџСЂРѕРµРєС‚ РїРµСЂРµРЅРµСЃРµРЅ РІ Р°СЂС…РёРІ."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def unarchive(self, request, pk=None):
        project = self.get_object()
        if request.user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if request.user.role == UserRole.TEACHER and project.supervisor_id != request.user.id:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        project.is_archived = False
        project.save(update_fields=["is_archived", "updated_at"])
        return Response({"detail": "РџСЂРѕРµРєС‚ РІРѕСЃСЃС‚Р°РЅРѕРІР»РµРЅ РёР· Р°СЂС…РёРІР°."}, status=status.HTTP_200_OK)

    @action(
        detail=False,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_cover(self, request):
        if request.user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "РќСѓР¶РЅРѕ РїРµСЂРµРґР°С‚СЊ С„Р°Р№Р»."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_uploaded_file(
                uploaded,
                policy=UploadPolicy(
                    allowed_extensions={".jpg", ".jpeg", ".png", ".webp", ".gif"},
                    max_size_bytes=MAX_IMAGE_UPLOAD_SIZE,
                    allow_images=True,
                ),
            )
        except FileValidationError as exc:
            return error_response(code="invalid_request", message=str(exc), http_status=status.HTTP_400_BAD_REQUEST)

        safe_name = sanitize_filename(uploaded.name, default_stem="project-cover")
        saved_path = default_storage.save(f"project_covers/{safe_name}", uploaded).replace("\\", "/")
        media_url = f"{settings.MEDIA_URL}{saved_path}"
        absolute_url = request.build_absolute_uri(media_url)

        return Response({"cover_image_url": absolute_url}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def portfolio(self, request):
        search = (request.query_params.get("search") or "").strip()
        date_from = parse_date(request.query_params.get("date_from") or "")
        date_to = parse_date(request.query_params.get("date_to") or "")
        limit = int(request.query_params.get("limit") or 20)
        limit = max(1, min(limit, 100))

        done_projects_filter = Q(projects__status=ProjectStatus.DONE)
        done_supervised_filter = Q(supervised_projects__status=ProjectStatus.DONE)
        if date_from:
            done_projects_filter &= Q(projects__end_date__gte=date_from)
            done_supervised_filter &= Q(supervised_projects__end_date__gte=date_from)
        if date_to:
            done_projects_filter &= Q(projects__end_date__lte=date_to)
            done_supervised_filter &= Q(supervised_projects__end_date__lte=date_to)

        student_queryset = User.objects.filter(role=UserRole.STUDENT)
        teacher_queryset = User.objects.filter(role=UserRole.TEACHER)
        if search:
            search_filter = (
                Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
                | Q(username__icontains=search)
                | Q(group_name__icontains=search)
            )
            student_queryset = student_queryset.filter(search_filter)
            teacher_queryset = teacher_queryset.filter(search_filter)

        top_students = list(
            student_queryset.annotate(completed_count=Count("projects", filter=done_projects_filter, distinct=True))
            .values("id", "username", "first_name", "last_name", "group_name", "completed_count")
            .order_by("-completed_count", "last_name", "first_name")[:limit]
        )
        top_teachers = list(
            teacher_queryset.annotate(
                completed_count=Count("supervised_projects", filter=done_supervised_filter, distinct=True)
            )
            .values("id", "username", "first_name", "last_name", "completed_count")
            .order_by("-completed_count", "last_name", "first_name")[:limit]
        )

        return Response(
            {
                "date_from": str(date_from) if date_from else None,
                "date_to": str(date_to) if date_to else None,
                "top_students": top_students,
                "top_teachers": top_teachers,
            }
        )

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def methodist_report(self, request):
        if request.user.role not in {UserRole.METHODIST, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        date_from = parse_date(request.query_params.get("date_from") or "")
        date_to = parse_date(request.query_params.get("date_to") or "")
        teacher_search = (request.query_params.get("teacher_search") or "").strip()
        limit = int(request.query_params.get("limit") or 20)
        limit = max(1, min(limit, 200))

        done_filter = Q(status=ProjectStatus.DONE)
        done_supervised_filter = Q(supervised_projects__status=ProjectStatus.DONE)
        if date_from:
            done_filter &= Q(end_date__gte=date_from)
            done_supervised_filter &= Q(supervised_projects__end_date__gte=date_from)
        if date_to:
            done_filter &= Q(end_date__lte=date_to)
            done_supervised_filter &= Q(supervised_projects__end_date__lte=date_to)

        completed_projects = Project.objects.filter(done_filter)
        completed_total = completed_projects.count()

        completed_by_type = list(
            completed_projects.values("type").annotate(total=Count("id")).order_by("-total", "type")
        )
        projects_by_status = list(
            Project.objects.values("status").annotate(total=Count("id")).order_by("-total", "status")
        )

        teacher_queryset = User.objects.filter(role=UserRole.TEACHER)
        if teacher_search:
            teacher_queryset = teacher_queryset.filter(
                Q(first_name__icontains=teacher_search)
                | Q(last_name__icontains=teacher_search)
                | Q(username__icontains=teacher_search)
            )
        teacher_workload = list(
            teacher_queryset.annotate(
                total_projects=Count("supervised_projects", distinct=True),
                completed_projects=Count("supervised_projects", filter=done_supervised_filter, distinct=True),
            )
            .values("id", "username", "first_name", "last_name", "total_projects", "completed_projects")
            .order_by("-total_projects", "-completed_projects", "last_name", "first_name")[:limit]
        )

        student_performance = list(
            User.objects.filter(role=UserRole.STUDENT)
            .annotate(
                completed_projects=Count("projects", filter=Q(projects__status=ProjectStatus.DONE), distinct=True),
            )
            .values("id", "username", "first_name", "last_name", "group_name", "completed_projects")
            .order_by("-completed_projects", "last_name", "first_name")[:limit]
        )

        avg_completed_by_teacher = 0
        if teacher_workload:
            avg_completed_by_teacher = round(
                sum(row["completed_projects"] for row in teacher_workload) / len(teacher_workload), 2
            )

        return Response(
            {
                "date_from": str(date_from) if date_from else None,
                "date_to": str(date_to) if date_to else None,
                "completed_total": completed_total,
                "completed_by_type": completed_by_type,
                "projects_by_status": projects_by_status,
                "teacher_workload": teacher_workload,
                "student_performance": student_performance,
                "avg_completed_by_teacher": avg_completed_by_teacher,
            }
        )

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def methodist_analytics(self, request):
        if request.user.role not in {UserRole.METHODIST, UserRole.CURATOR, UserRole.ADMIN, UserRole.TEACHER, UserRole.STUDENT}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        date_from = parse_date(request.query_params.get("date_from") or "")
        date_to = parse_date(request.query_params.get("date_to") or "")
        project_type = (request.query_params.get("project_type") or "").strip()
        supervisor_id = (request.query_params.get("supervisor") or "").strip()
        selected_status = (request.query_params.get("selected_status") or "").strip()
        selected_type = (request.query_params.get("selected_type") or "").strip()
        selected_month = (request.query_params.get("selected_month") or "").strip()

        queryset = Project.objects.select_related("supervisor")
        analytics_scope = "platform"
        if request.user.role == UserRole.TEACHER:
            analytics_scope = "teacher_personal"
            queryset = queryset.filter(supervisor_id=request.user.id)
        elif request.user.role == UserRole.STUDENT:
            analytics_scope = "student_personal"
            queryset = queryset.filter(participants__id=request.user.id).distinct()
        if date_from:
            queryset = queryset.filter(created_at__date__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__date__lte=date_to)
        if project_type:
            queryset = queryset.filter(type=project_type)
        if supervisor_id.isdigit() and analytics_scope == "platform":
            queryset = queryset.filter(supervisor_id=int(supervisor_id))
        if selected_status:
            queryset = queryset.filter(status=selected_status)
        if selected_type:
            queryset = queryset.filter(type=selected_type)
        if selected_month:
            try:
                year, month = selected_month.split("-")
                month_start = _month_start(int(year), int(month))
                next_month = _shift_month(month_start, 1)
                queryset = queryset.filter(created_at__date__gte=month_start, created_at__date__lt=next_month)
            except Exception:
                return error_response(
                    code="invalid_month_filter",
                    message="Некорректный формат selected_month. Ожидается YYYY-MM.",
                    http_status=status.HTTP_400_BAD_REQUEST,
                )

        raw_status_counts = list(queryset.values("status").annotate(total=Count("id")))
        raw_type_counts = list(queryset.values("type").annotate(total=Count("id")))
        status_map = {row["status"]: int(row["total"]) for row in raw_status_counts}
        type_map = {row["type"]: int(row["total"]) for row in raw_type_counts}
        status_counts = [{"status": key, "total": status_map.get(key, 0)} for key, _ in ProjectStatus.choices]
        type_counts = [{"type": key, "total": type_map.get(key, 0)} for key, _ in ProjectType.choices]

        now_local = timezone.localdate()
        current_month_start = _month_start(now_local.year, now_local.month)
        first_month_start = _shift_month(current_month_start, -11)
        monthly_raw = (
            queryset.filter(created_at__date__gte=first_month_start, created_at__date__lt=_shift_month(current_month_start, 1))
            .annotate(month=TruncMonth("created_at"))
            .values("month")
            .annotate(total=Count("id"))
            .order_by("month")
        )
        monthly_map = {row["month"].strftime("%Y-%m"): row["total"] for row in monthly_raw if row.get("month")}
        monthly_counts = []
        for offset in range(12):
            month_start = _shift_month(first_month_start, offset)
            key = month_start.strftime("%Y-%m")
            monthly_counts.append({"month": key, "total": int(monthly_map.get(key, 0))})

        projects = list(
            queryset.order_by("-created_at")
            .values(
                "id",
                "title",
                "type",
                "status",
                "created_at",
                "start_date",
                "end_date",
                "supervisor_id",
                "supervisor__first_name",
                "supervisor__last_name",
                "supervisor__username",
            )[:300]
        )
        for row in projects:
            supervisor_name = f"{row.get('supervisor__last_name') or ''} {row.get('supervisor__first_name') or ''}".strip()
            row["supervisor_name"] = supervisor_name or row.get("supervisor__username") or "-"

        supervisors = list(
            User.objects.filter(role=UserRole.TEACHER, is_active=True)
            .values("id", "first_name", "last_name", "username")
            .order_by("last_name", "first_name", "username")
        )
        for teacher in supervisors:
            full_name = f"{teacher.get('last_name') or ''} {teacher.get('first_name') or ''}".strip()
            teacher["full_name"] = full_name or teacher.get("username")

        return Response(
            {
                "filters": {
                    "date_from": str(date_from) if date_from else None,
                    "date_to": str(date_to) if date_to else None,
                    "project_type": project_type or None,
                    "supervisor": int(supervisor_id) if supervisor_id.isdigit() else None,
                    "selected_status": selected_status or None,
                    "selected_type": selected_type or None,
                    "selected_month": selected_month or None,
                },
                "scope": analytics_scope,
                "charts": {
                    "status_counts": status_counts,
                    "type_counts": type_counts,
                    "monthly_counts": monthly_counts,
                },
                "table": {
                    "projects": projects,
                    "count": len(projects),
                },
                "supervisors": supervisors,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def teacher_workload(self, request):
        if request.user.role not in {UserRole.METHODIST, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        now_local = timezone.localdate()
        year = int(request.query_params.get("year") or now_local.year)
        month = int(request.query_params.get("month") or now_local.month)
        low_max = int(request.query_params.get("low_max") or 0)
        normal_max = int(request.query_params.get("normal_max") or 1)
        high_max = int(request.query_params.get("high_max") or 2)
        export = (request.query_params.get("export") or "").strip().lower()

        month_start, month_end = _month_range(year, month)
        prev_1 = _shift_month(month_start, -1)
        prev_2 = _shift_month(month_start, -2)
        prev_3 = _shift_month(month_start, -3)

        active_statuses = [ProjectStatus.PLANNED, ProjectStatus.IN_PROGRESS, ProjectStatus.REVIEW]
        teachers = list(
            User.objects.filter(role=UserRole.TEACHER, is_active=True)
            .values("id", "first_name", "last_name", "username")
            .order_by("last_name", "first_name", "username")
        )

        teacher_ids = [teacher["id"] for teacher in teachers]
        active_map = {
            item["supervisor_id"]: item["count"]
            for item in Project.objects.filter(supervisor_id__in=teacher_ids, status__in=active_statuses)
            .values("supervisor_id")
            .annotate(count=Count("id"))
        }
        review_map = {
            item["project__supervisor_id"]: item["count"]
            for item in ProjectStage.objects.filter(
                project__supervisor_id__in=teacher_ids,
                status=StageStatus.SUBMITTED,
            )
            .values("project__supervisor_id")
            .annotate(count=Count("id"))
        }
        m1_map = {
            item["supervisor_id"]: item["count"]
            for item in Project.objects.filter(
                supervisor_id__in=teacher_ids,
                created_at__date__gte=prev_1,
                created_at__date__lt=month_start,
            )
            .values("supervisor_id")
            .annotate(count=Count("id"))
        }
        m2_map = {
            item["supervisor_id"]: item["count"]
            for item in Project.objects.filter(
                supervisor_id__in=teacher_ids,
                created_at__date__gte=prev_2,
                created_at__date__lt=prev_1,
            )
            .values("supervisor_id")
            .annotate(count=Count("id"))
        }
        m3_map = {
            item["supervisor_id"]: item["count"]
            for item in Project.objects.filter(
                supervisor_id__in=teacher_ids,
                created_at__date__gte=prev_3,
                created_at__date__lt=prev_2,
            )
            .values("supervisor_id")
            .annotate(count=Count("id"))
        }
        done_last_map = {
            item["supervisor_id"]: item["count"]
            for item in Project.objects.filter(
                supervisor_id__in=teacher_ids,
                status=ProjectStatus.DONE,
                updated_at__date__gte=prev_1,
                updated_at__date__lt=month_start,
            )
            .values("supervisor_id")
            .annotate(count=Count("id"))
        }

        rows = []
        for teacher in teachers:
            teacher_id = teacher["id"]
            active_projects = active_map.get(teacher_id, 0)
            review_stages = review_map.get(teacher_id, 0)

            m1_new = m1_map.get(teacher_id, 0)
            m2_new = m2_map.get(teacher_id, 0)
            m3_new = m3_map.get(teacher_id, 0)
            avg_new = round((m1_new + m2_new + m3_new) / 3, 1)
            done_last_month = done_last_map.get(teacher_id, 0)
            forecast_next = max(0, int(round(active_projects + avg_new - done_last_month)))
            delta = forecast_next - active_projects

            zone = _load_zone(active_projects, low_max=low_max, normal_max=normal_max, high_max=high_max)
            if zone in {"high", "critical"}:
                action = "reduce"
            elif zone == "low":
                action = "increase"
            else:
                action = "keep"

            full_name = f"{teacher.get('last_name') or ''} {teacher.get('first_name') or ''}".strip() or teacher.get("username")
            rows.append(
                {
                    "teacher_id": teacher_id,
                    "teacher_name": full_name,
                    "active_projects": active_projects,
                    "review_stages": review_stages,
                    "forecast_next_month": forecast_next,
                    "forecast_delta": delta,
                    "zone": zone,
                    "action": action,
                }
            )

        total_teachers = len(rows)
        critical_count = sum(1 for row in rows if row["zone"] == "critical")
        low_count = sum(1 for row in rows if row["zone"] == "low")
        avg_load = round(sum(row["active_projects"] for row in rows) / total_teachers, 2) if total_teachers else 0

        history = []
        for offset in range(-5, 1):
            m_start = _shift_month(month_start, offset)
            m_end = _shift_month(m_start, 1) - timedelta(days=1)
            month_counts_map = {
                item["supervisor_id"]: item["count"]
                for item in Project.objects.filter(
                    supervisor_id__in=teacher_ids,
                    status__in=active_statuses,
                    created_at__date__lte=m_end,
                )
                .values("supervisor_id")
                .annotate(count=Count("id"))
            }
            teacher_month_rows = [month_counts_map.get(teacher_id, 0) for teacher_id in teacher_ids]
            avg_month = round(sum(teacher_month_rows) / len(teacher_month_rows), 2) if teacher_month_rows else 0
            overloaded_month = sum(1 for value in teacher_month_rows if value > high_max)
            history.append(
                {
                    "month": m_start.strftime("%Y-%m"),
                    "avg_load": avg_month,
                    "overloaded_count": overloaded_month,
                }
            )

        if export == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Нагрузка преподавателей"
            title = f"ОТЧЕТ ПО НАГРУЗКЕ ПРЕПОДАВАТЕЛЕЙ ЗА {year}-{month:02d}"
            sheet.merge_cells("A1:G1")
            sheet["A1"] = title
            sheet["A1"].font = Font(name="Times New Roman", size=14, bold=True)
            sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 24

            sheet["A3"] = "Сводка"
            sheet["A3"].font = Font(name="Times New Roman", size=12, bold=True)
            sheet["A4"] = f"Всего преподавателей: {total_teachers}"
            sheet["A5"] = f"С перегрузкой (>{high_max} проектов): {critical_count}"
            sheet["A6"] = f"С низкой загрузкой (до {low_max} проектов): {low_count}"
            sheet["A7"] = f"Средняя нагрузка: {avg_load}"
            for row_idx in range(4, 8):
                sheet[f"A{row_idx}"].font = Font(name="Times New Roman", size=11)

            header_row = 9
            sheet.append([])
            sheet.append(
                [
                    "Преподаватель",
                    "Активных проектов",
                    "На проверке (этапов)",
                    "Прогноз на след. месяц",
                    "Изменение",
                    "Зона",
                    "Рекомендация",
                ]
            )
            header_fill = PatternFill("solid", fgColor="D9E1F2")
            thin = Side(style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col in "ABCDEFG":
                cell = sheet[f"{col}{header_row}"]
                cell.font = Font(name="Times New Roman", size=11, bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border

            zone_labels = {
                "low": "Низкая",
                "normal": "Нормальная",
                "high": "Высокая",
                "critical": "Критическая",
            }
            action_labels = {
                "reduce": "СНИЗИТЬ",
                "increase": "ПОВЫСИТЬ",
                "keep": "БЕЗ ИЗМЕНЕНИЙ",
            }
            for row in rows:
                sheet.append(
                    [
                        row["teacher_name"],
                        row["active_projects"],
                        row["review_stages"],
                        row["forecast_next_month"],
                        row["forecast_delta"],
                        zone_labels.get(row["zone"], row["zone"]),
                        action_labels.get(row["action"], row["action"]),
                    ]
                )
            start_data = header_row + 1
            end_data = header_row + len(rows)
            for row_idx in range(start_data, end_data + 1):
                for col in "ABCDEFG":
                    cell = sheet[f"{col}{row_idx}"]
                    cell.font = Font(name="Times New Roman", size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = border
                zone_value = sheet[f"F{row_idx}"].value
                if zone_value == "Критическая":
                    fill = PatternFill("solid", fgColor="F8CBAD")
                elif zone_value == "Высокая":
                    fill = PatternFill("solid", fgColor="FCE4D6")
                elif zone_value == "Низкая":
                    fill = PatternFill("solid", fgColor="FFF2CC")
                else:
                    fill = PatternFill("solid", fgColor="E2F0D9")
                sheet[f"F{row_idx}"].fill = fill
                sheet[f"G{row_idx}"].fill = fill

            column_widths = {
                "A": 34,
                "B": 20,
                "C": 24,
                "D": 22,
                "E": 14,
                "F": 16,
                "G": 18,
            }
            for col, width in column_widths.items():
                sheet.column_dimensions[col].width = width

            output = BytesIO()
            workbook.save(output)
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="teacher_workload_{year}_{month:02d}.xlsx"'
            return response

        return Response(
            {
                "period": {"year": year, "month": month, "month_start": str(month_start), "month_end": str(month_end)},
                "thresholds": {"low_max": low_max, "normal_max": normal_max, "high_max": high_max},
                "summary": {
                    "total_teachers": total_teachers,
                    "critical_overloaded": critical_count,
                    "low_load": low_count,
                    "avg_load": avg_load,
                },
                "rows": rows,
                "history": history,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def teacher_workload_projects(self, request):
        if request.user.role not in {UserRole.METHODIST, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        teacher_id = int(request.query_params.get("teacher_id") or 0)
        if not teacher_id:
            return error_response(code="invalid_teacher_id", message="Укажите корректный teacher_id.", http_status=status.HTTP_400_BAD_REQUEST)

        active_statuses = [ProjectStatus.PLANNED, ProjectStatus.IN_PROGRESS, ProjectStatus.REVIEW]
        projects = list(
            Project.objects.filter(supervisor_id=teacher_id, status__in=active_statuses)
            .values("id", "title", "status", "type")
            .order_by("-updated_at", "-created_at")
        )
        return Response({"projects": projects}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def teacher_workload_reassign(self, request):
        if request.user.role not in {UserRole.METHODIST, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        project_id = int(request.data.get("project_id") or 0)
        target_teacher_id = int(request.data.get("target_teacher_id") or 0)
        if not project_id or not target_teacher_id:
            return error_response(
                code="invalid_payload",
                message="Передайте project_id и target_teacher_id.",
                http_status=status.HTTP_400_BAD_REQUEST,
            )

        project = Project.objects.filter(id=project_id).select_related("supervisor").first()
        if not project:
            return error_response(code="project_not_found", message="Проект не найден.", http_status=status.HTTP_404_NOT_FOUND)
        if project.status in {ProjectStatus.DONE, ProjectStatus.CANCELLED}:
            return error_response(code="project_closed", message="Нельзя переназначить завершенный/отмененный проект.", http_status=status.HTTP_400_BAD_REQUEST)

        target = User.objects.filter(id=target_teacher_id, role=UserRole.TEACHER, is_active=True).first()
        if not target:
            return error_response(code="teacher_not_found", message="Выбранный преподаватель недоступен.", http_status=status.HTTP_404_NOT_FOUND)

        previous_supervisor = project.supervisor
        if previous_supervisor_id := getattr(previous_supervisor, "id", None):
            if previous_supervisor_id == target.id:
                return error_response(code="same_supervisor", message="Проект уже закреплен за этим преподавателем.", http_status=status.HTTP_400_BAD_REQUEST)

        project.supervisor = target
        project.save(update_fields=["supervisor", "updated_at"])

        create_notifications(
            [target],
            actor=request.user,
            project=project,
            type=NotificationType.PROJECT_ASSIGNED,
            title="Назначено руководство проектом",
            message=f"Вам назначен проект «{project.title}».",
        )
        if previous_supervisor and previous_supervisor.id != target.id:
            create_notifications(
                [previous_supervisor],
                actor=request.user,
                project=project,
                type=NotificationType.STAGE_UPDATED,
                title="Проект передан другому преподавателю",
                message=f"Проект «{project.title}» передан преподавателю {target.full_name or target.username}.",
            )

        return Response(
            {
                "detail": "Руководитель проекта успешно изменен.",
                "project_id": project.id,
                "new_supervisor_id": target.id,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def teacher_deadlines(self, request):
        user = request.user
        if user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        days = int(request.query_params.get("days") or 14)
        days = max(1, min(days, 90))
        date_to = timezone.localdate() + timedelta(days=days)

        stages_qs = (
            ProjectStage.objects.select_related("project")
            .filter(deadline__isnull=False, deadline__lte=date_to, status__in=[StageStatus.OPEN, StageStatus.SUBMITTED])
            .order_by("deadline", "project_id", "order")
        )
        if user.role == UserRole.TEACHER:
            stages_qs = stages_qs.filter(project__supervisor_id=user.id)

        stages = list(
            stages_qs.values(
                "id",
                "title",
                "status",
                "deadline",
                "order",
                "project_id",
                "project__title",
                "project__status",
            )
        )
        return Response({"days": days, "count": len(stages), "items": stages})

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def request_publish(self, request, pk=None):
        project = self.get_object()
        user = request.user
        if user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if user.role == UserRole.TEACHER and project.supervisor_id != user.id:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if project.status != ProjectStatus.DONE:
            return Response({"detail": "РџСЂРѕРµРєС‚ РґРѕР»Р¶РµРЅ Р±С‹С‚СЊ РІ СЃС‚Р°С‚СѓСЃРµ В«Р—Р°РІРµСЂС€РµРЅВ»."}, status=status.HTTP_400_BAD_REQUEST)

        recipients = User.objects.filter(role__in=[UserRole.CURATOR, UserRole.ADMIN], is_active=True)
        create_notifications(
            recipients,
            actor=user,
            project=project,
            type=NotificationType.PROJECT_PUBLISH_REQUEST,
            title="Р—Р°РїСЂРѕСЃ РЅР° РїСѓР±Р»РёРєР°С†РёСЋ РїСЂРѕРµРєС‚Р°",
            message=f"{user.full_name or user.username} РїСЂРѕСЃРёС‚ РѕРїСѓР±Р»РёРєРѕРІР°С‚СЊ РїСЂРѕРµРєС‚ В«{project.title}В».",
        )
        return Response({"detail": "Р—Р°РїСЂРѕСЃ РЅР° РїСѓР±Р»РёРєР°С†РёСЋ РѕС‚РїСЂР°РІР»РµРЅ РєСѓСЂР°С‚РѕСЂСѓ."}, status=status.HTTP_200_OK)


class ProjectStageViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectStageSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ("project", "status")
    ordering_fields = ("order", "deadline", "updated_at")

    def get_queryset(self):
        return ProjectStage.objects.select_related("project", "updated_by")

    def perform_create(self, serializer):
        project = serializer.validated_data["project"]
        user = self.request.user
        if user.role in {UserRole.CURATOR, UserRole.ADMIN}:
            stage = serializer.save(updated_by=user)
            create_notifications(
                project.participants.all(),
                actor=user,
                project=project,
                stage=stage,
                type=NotificationType.STAGE_CREATED,
                title="РЎРѕР·РґР°РЅ РЅРѕРІС‹Р№ СЌС‚Р°Рї",
                message=f"Р­С‚Р°Рї В«{stage.title}В» РІ РїСЂРѕРµРєС‚Рµ В«{project.title}В».",
            )
            return
        if user.role == UserRole.TEACHER and project.supervisor_id == user.id:
            stage = serializer.save(updated_by=user)
            create_notifications(
                project.participants.all(),
                actor=user,
                project=project,
                stage=stage,
                type=NotificationType.STAGE_CREATED,
                title="РЎРѕР·РґР°РЅ РЅРѕРІС‹Р№ СЌС‚Р°Рї",
                message=f"Р­С‚Р°Рї В«{stage.title}В» РІ РїСЂРѕРµРєС‚Рµ В«{project.title}В».",
            )
            return
        raise permissions.PermissionDenied(MESSAGES["permission_denied"])

    def perform_update(self, serializer):
        serializer.save(updated_by=self.request.user)

    def update(self, request, *args, **kwargs):
        stage = self.get_object()
        old_status = stage.status
        user = request.user
        # РћРґРёРЅ РјР°СЂС€СЂСѓС‚ РѕР±СЃР»СѓР¶РёРІР°РµС‚ С‚СЂРё СЂРѕР»Рё СЃ СЂР°Р·РЅС‹РјРё РїСЂР°РІР°РјРё РёР·РјРµРЅРµРЅРёСЏ.
        if user.role in {UserRole.CURATOR, UserRole.ADMIN}:
            response = super().update(request, *args, **kwargs)
            stage.refresh_from_db()
            if stage.status != old_status and stage.status in {StageStatus.APPROVED, StageStatus.CHANGES_REQUESTED}:
                recipients = stage.project.participants.filter(role=UserRole.STUDENT)
                create_notifications(
                    recipients,
                    actor=user,
                    project=stage.project,
                    stage=stage,
                    type=NotificationType.STAGE_REVIEWED,
                    title="Р­С‚Р°Рї РїСЂРѕРІРµСЂРµРЅ",
                    message=f"Р­С‚Р°Рї В«{stage.title}В»: СЃС‚Р°С‚СѓСЃ РёР·РјРµРЅРµРЅ РЅР° {stage.status}.",
                )
            return response
        if user.role == UserRole.TEACHER and stage.project.supervisor_id == user.id:
            response = super().update(request, *args, **kwargs)
            stage.refresh_from_db()
            if stage.status != old_status and stage.status in {StageStatus.APPROVED, StageStatus.CHANGES_REQUESTED}:
                recipients = stage.project.participants.filter(role=UserRole.STUDENT)
                create_notifications(
                    recipients,
                    actor=user,
                    project=stage.project,
                    stage=stage,
                    type=NotificationType.STAGE_REVIEWED,
                    title="Р­С‚Р°Рї РїСЂРѕРІРµСЂРµРЅ",
                    message=f"Р­С‚Р°Рї В«{stage.title}В»: СЃС‚Р°С‚СѓСЃ РёР·РјРµРЅРµРЅ РЅР° {stage.status}.",
                )
            return response
        if user.role == UserRole.STUDENT and stage.project.participants.filter(id=user.id).exists():
            allowed_fields = {"student_report", "status"}
            if any(field not in allowed_fields for field in request.data.keys()):
                return Response(
                    {"detail": "РЎС‚СѓРґРµРЅС‚ РјРѕР¶РµС‚ РёР·РјРµРЅСЏС‚СЊ С‚РѕР»СЊРєРѕ РѕС‚С‡РµС‚ Рё СЃС‚Р°С‚СѓСЃ СЌС‚Р°РїР°."},
                    status=status.HTTP_403_FORBIDDEN,
                )
            stage_status = request.data.get("status")
            if stage_status and stage_status not in {StageStatus.OPEN, StageStatus.SUBMITTED}:
                return Response(
                    {"detail": "РЎС‚СѓРґРµРЅС‚ РјРѕР¶РµС‚ СѓСЃС‚Р°РЅРѕРІРёС‚СЊ С‚РѕР»СЊРєРѕ СЃС‚Р°С‚СѓСЃС‹ В«РћС‚РєСЂС‹С‚В» РёР»Рё В«РЎРґР°РЅ РЅР° РїСЂРѕРІРµСЂРєСѓВ»."},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            response = super().update(request, *args, **kwargs)
            stage.refresh_from_db()
            if stage.status == StageStatus.SUBMITTED:
                curator_admins = User.objects.filter(role__in=[UserRole.CURATOR, UserRole.ADMIN])
                recipients = list(curator_admins) + [stage.project.supervisor]
                create_notifications(
                    recipients,
                    actor=user,
                    project=stage.project,
                    stage=stage,
                    type=NotificationType.STAGE_SUBMITTED,
                    title="Р­С‚Р°Рї РѕС‚РїСЂР°РІР»РµРЅ РЅР° РїСЂРѕРІРµСЂРєСѓ",
                    message=f"{user.full_name or user.username} РѕС‚РїСЂР°РІРёР» СЌС‚Р°Рї В«{stage.title}В».",
                )
            return response
        return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

    def destroy(self, request, *args, **kwargs):
        stage = self.get_object()
        user = request.user
        if user.role in {UserRole.CURATOR, UserRole.ADMIN}:
            return super().destroy(request, *args, **kwargs)
        if user.role == UserRole.TEACHER and stage.project.supervisor_id == user.id:
            return super().destroy(request, *args, **kwargs)
        return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)


class ProjectCommentViewSet(
    mixins.CreateModelMixin,
    mixins.ListModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    serializer_class = ProjectCommentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ("project", "stage", "is_approved")
    search_fields = ("text", "author__first_name", "author__last_name")
    ordering_fields = ("created_at",)

    def get_queryset(self):
        queryset = ProjectComment.objects.select_related("project", "author", "stage")
        if self.request.user.role in {UserRole.CURATOR, UserRole.ADMIN}:
            return queryset
        return queryset

    def perform_create(self, serializer):
        comment = serializer.save(author=self.request.user, is_approved=True)
        recipients = comment.project.participants.exclude(id=self.request.user.id)
        if comment.project.supervisor_id != self.request.user.id:
            recipients = (recipients | User.objects.filter(id=comment.project.supervisor_id)).distinct()
        recipient_ids = list(recipients.values_list("id", flat=True))
        message = f"{self.request.user.full_name or self.request.user.username}: {comment.text[:200]}"
        try:
            dispatch_comment_notifications_task.delay(
                recipient_ids=recipient_ids,
                actor_id=self.request.user.id,
                project_id=comment.project_id,
                stage_id=comment.stage_id,
                message=message,
            )
        except Exception:
            logger.exception("Celery unavailable for comment notifications; fallback to sync send")
            create_notifications(
                User.objects.filter(id__in=recipient_ids),
                actor=self.request.user,
                project=comment.project,
                stage=comment.stage,
                type=NotificationType.COMMENT_PENDING,
                title="Комментарий опубликован",
                message=message,
            )

    @action(detail=True, methods=["post"], permission_classes=[IsCuratorOrReadOnly])
    def approve(self, request, pk=None):
        comment = self.get_object()
        comment.is_approved = True
        comment.save(update_fields=["is_approved"])
        return Response({"detail": "РљРѕРјРјРµРЅС‚Р°СЂРёР№ РїРѕРґС‚РІРµСЂР¶РґРµРЅ."})


class ProjectLikeViewSet(mixins.ListModelMixin, viewsets.GenericViewSet):
    serializer_class = ProjectLikeSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ("project", "user")
    ordering_fields = ("created_at",)

    def get_queryset(self):
        return ProjectLike.objects.select_related("project", "user")


class TeamManageViewSet(viewsets.ModelViewSet):
    serializer_class = TeamManageSerializer
    permission_classes = [permissions.IsAuthenticated]
    search_fields = ("name", "group_name")
    ordering_fields = ("name", "created_at", "members_total")

    def get_queryset(self):
        queryset = (
            Team.objects.exclude(name__startswith="group:")
            .prefetch_related("members")
            .select_related("supervisor")
            .annotate(members_total=Count("members", distinct=True))
        )

        has_photo = (self.request.query_params.get("has_photo") or "").strip().lower()
        members_min = (self.request.query_params.get("members_min") or "").strip()
        members_max = (self.request.query_params.get("members_max") or "").strip()

        if has_photo == "true":
            queryset = queryset.exclude(photo_url="")
        elif has_photo == "false":
            queryset = queryset.filter(photo_url="")

        if members_min.isdigit():
            queryset = queryset.filter(members_total__gte=int(members_min))
        if members_max.isdigit():
            queryset = queryset.filter(members_total__lte=int(members_max))

        if self.request.user.role == UserRole.TEACHER:
            return queryset.filter(supervisor_id=self.request.user.id)
        return queryset

    def perform_create(self, serializer):
        user = self.request.user
        if user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            raise permissions.PermissionDenied(MESSAGES["permission_denied"])
        serializer.save(supervisor=user)

    def perform_update(self, serializer):
        user = self.request.user
        team = self.get_object()
        if user.role == UserRole.TEACHER and team.supervisor_id != user.id:
            raise permissions.PermissionDenied(MESSAGES["permission_denied"])
        if user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            raise permissions.PermissionDenied(MESSAGES["permission_denied"])
        serializer.save()

    def perform_destroy(self, instance):
        user = self.request.user
        if user.role == UserRole.TEACHER and instance.supervisor_id != user.id:
            raise permissions.PermissionDenied(MESSAGES["permission_denied"])
        if user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            raise permissions.PermissionDenied(MESSAGES["permission_denied"])
        instance.delete()

    @action(
        detail=False,
        methods=["post"],
        permission_classes=[permissions.IsAuthenticated],
        parser_classes=[MultiPartParser, FormParser],
    )
    def upload_photo(self, request):
        if request.user.role not in {UserRole.TEACHER, UserRole.CURATOR, UserRole.ADMIN}:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)

        uploaded = request.FILES.get("file")
        if not uploaded:
            return Response({"detail": "РќСѓР¶РЅРѕ РїРµСЂРµРґР°С‚СЊ С„Р°Р№Р»."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            validate_uploaded_file(
                uploaded,
                policy=UploadPolicy(
                    allowed_extensions={".jpg", ".jpeg", ".png", ".webp", ".gif"},
                    max_size_bytes=MAX_IMAGE_UPLOAD_SIZE,
                    allow_images=True,
                ),
            )
        except FileValidationError as exc:
            return error_response(code="invalid_request", message=str(exc), http_status=status.HTTP_400_BAD_REQUEST)

        safe_name = sanitize_filename(uploaded.name, default_stem="team-photo")
        saved_path = default_storage.save(f"team_photos/{safe_name}", uploaded).replace("\\", "/")
        media_url = f"{settings.MEDIA_URL}{saved_path}"
        absolute_url = request.build_absolute_uri(media_url)
        return Response({"photo_url": absolute_url}, status=status.HTTP_201_CREATED)


class ProjectSupervisorInviteViewSet(viewsets.ModelViewSet):
    serializer_class = ProjectSupervisorInviteSerializer
    permission_classes = [permissions.IsAuthenticated]
    http_method_names = ["get", "post", "patch", "head", "options"]
    ordering_fields = ("created_at", "status")

    def get_queryset(self):
        queryset = ProjectSupervisorInvite.objects.select_related("project", "student", "teacher")
        user = self.request.user
        if user.role == UserRole.TEACHER:
            return queryset.filter(teacher=user)
        if user.role == UserRole.STUDENT:
            return queryset.filter(student=user)
        if user.role in {UserRole.CURATOR, UserRole.ADMIN}:
            return queryset
        return queryset.none()

    def perform_create(self, serializer):
        invite = serializer.save(student=self.request.user)
        create_notifications(
            [invite.teacher],
            actor=self.request.user,
            project=invite.project,
            type=NotificationType.SUPERVISOR_INVITED,
            title="Р’Р°СЃ РїСЂРёРіР»Р°СЃРёР»Рё СЂСѓРєРѕРІРѕРґРёС‚РµР»РµРј РїСЂРѕРµРєС‚Р°",
            message=f"РџСЂРѕРµРєС‚: {invite.project.title}",
        )

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def accept(self, request, pk=None):
        invite = self.get_object()
        if request.user.id != invite.teacher_id:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if invite.status != SupervisorInviteStatus.PENDING:
            return Response({"detail": "РџСЂРёРіР»Р°С€РµРЅРёРµ СѓР¶Рµ РѕР±СЂР°Р±РѕС‚Р°РЅРѕ."}, status=status.HTTP_400_BAD_REQUEST)

        invite.status = SupervisorInviteStatus.ACCEPTED
        invite.responded_at = timezone.now()
        invite.save(update_fields=["status", "responded_at"])

        project = invite.project
        project.supervisor = request.user
        project.save(update_fields=["supervisor", "updated_at"])

        create_notifications(
            [invite.student],
            actor=request.user,
            project=project,
            type=NotificationType.SUPERVISOR_INVITE_ACCEPTED,
            title="РџСЂРµРїРѕРґР°РІР°С‚РµР»СЊ РїСЂРёРЅСЏР» РїСЂРёРіР»Р°С€РµРЅРёРµ",
            message=f"{request.user.full_name or request.user.username} СЃС‚Р°Р» СЂСѓРєРѕРІРѕРґРёС‚РµР»РµРј РїСЂРѕРµРєС‚Р° {project.title}",
        )

        return Response({"detail": "РџСЂРёРіР»Р°С€РµРЅРёРµ РїСЂРёРЅСЏС‚Рѕ."})

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def decline(self, request, pk=None):
        invite = self.get_object()
        if request.user.id != invite.teacher_id:
            return error_response(code="permission_denied", message=MESSAGES["permission_denied"], http_status=status.HTTP_403_FORBIDDEN)
        if invite.status != SupervisorInviteStatus.PENDING:
            return Response({"detail": "РџСЂРёРіР»Р°С€РµРЅРёРµ СѓР¶Рµ РѕР±СЂР°Р±РѕС‚Р°РЅРѕ."}, status=status.HTTP_400_BAD_REQUEST)

        invite.status = SupervisorInviteStatus.DECLINED
        invite.responded_at = timezone.now()
        invite.save(update_fields=["status", "responded_at"])

        create_notifications(
            [invite.student],
            actor=request.user,
            project=invite.project,
            type=NotificationType.SUPERVISOR_INVITE_DECLINED,
            title="РџСЂРµРїРѕРґР°РІР°С‚РµР»СЊ РѕС‚РєР»РѕРЅРёР» РїСЂРёРіР»Р°С€РµРЅРёРµ",
            message=f"{request.user.full_name or request.user.username} РѕС‚РєР»РѕРЅРёР» РїСЂРёРіР»Р°С€РµРЅРёРµ РІ РїСЂРѕРµРєС‚ {invite.project.title}",
        )
        return Response({"detail": "РџСЂРёРіР»Р°С€РµРЅРёРµ РѕС‚РєР»РѕРЅРµРЅРѕ."})








